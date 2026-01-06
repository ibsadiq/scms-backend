"""
Celery tasks for user operations.

Handles asynchronous operations like:
- Bulk teacher uploads
- Bulk parent uploads
- Email sending for invitations
"""
from celery import shared_task
from django.db import transaction
from django.core.mail import send_mail
from django.conf import settings
import openpyxl
from io import BytesIO

from academic.models import Teacher, Parent, Student
from users.models import CustomUser as User


@shared_task(bind=True, name='users.bulk_upload_teachers')
def bulk_upload_teachers_task(self, file_content):
    """
    Async task for bulk uploading teachers from Excel file.

    Args:
        self: Celery task instance
        file_content: Binary file content

    Returns:
        dict: Summary of upload results
    """
    results = {
        'total_rows': 0,
        'created': 0,
        'failed': 0,
        'errors': []
    }

    try:
        workbook = openpyxl.load_workbook(BytesIO(file_content))
        sheet = workbook.active

        columns = [
            "first_name",
            "last_name",
            "email",
            "phone_number",
            "date_of_birth",
            "gender",
            "address",
            "national_id",
            "tin_number",
            "qualification",
        ]

        teachers_to_create = []

        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            results['total_rows'] += 1

            # Update progress
            self.update_state(
                state='PROGRESS',
                meta={'current': i - 1, 'total': sheet.max_row - 1, 'status': f'Processing row {i}'}
            )

            teacher_data = dict(zip(columns, row))

            try:
                # Check if teacher already exists
                if Teacher.objects.filter(email=teacher_data["email"]).exists():
                    raise ValueError(f"Row {i}: Teacher with email '{teacher_data['email']}' already exists")

                # Check for duplicate NIN
                if teacher_data.get("national_id"):
                    if Teacher.objects.filter(national_id=teacher_data["national_id"]).exists():
                        raise ValueError(f"Row {i}: Teacher with NIN '{teacher_data['national_id']}' already exists")

                # Check for duplicate TIN
                if teacher_data.get("tin_number"):
                    if Teacher.objects.filter(tin_number=teacher_data["tin_number"]).exists():
                        raise ValueError(f"Row {i}: Teacher with TIN '{teacher_data['tin_number']}' already exists")

                teacher = Teacher(
                    first_name=teacher_data["first_name"],
                    last_name=teacher_data["last_name"],
                    email=teacher_data["email"],
                    phone_number=teacher_data.get("phone_number"),
                    date_of_birth=teacher_data.get("date_of_birth"),
                    gender=teacher_data["gender"].upper()[0],
                    address=teacher_data.get("address"),
                    national_id=teacher_data.get("national_id"),
                    tin_number=teacher_data.get("tin_number"),
                    qualification=teacher_data.get("qualification"),
                )

                teachers_to_create.append(teacher)

            except Exception as e:
                results['failed'] += 1
                results['errors'].append(str(e))

        # Bulk create
        with transaction.atomic():
            for teacher in teachers_to_create:
                teacher.save()
                results['created'] += 1

        return results

    except Exception as e:
        results['errors'].append(f"Critical error: {str(e)}")
        return results


@shared_task(bind=True, name='users.bulk_upload_parents')
def bulk_upload_parents_task(self, file_content):
    """
    Async task for bulk uploading parents from Excel file.

    Args:
        self: Celery task instance
        file_content: Binary file content

    Returns:
        dict: Summary of upload results
    """
    results = {
        'total_rows': 0,
        'created': 0,
        'failed': 0,
        'errors': []
    }

    try:
        workbook = openpyxl.load_workbook(BytesIO(file_content))
        sheet = workbook.active

        columns = [
            "first_name",
            "last_name",
            "email",
            "phone_number",
            "address",
            "national_id",
            "relationship",
            "occupation",
            "student_admission_number",
        ]

        parents_to_create = []

        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            results['total_rows'] += 1

            # Update progress
            self.update_state(
                state='PROGRESS',
                meta={'current': i - 1, 'total': sheet.max_row - 1, 'status': f'Processing row {i}'}
            )

            parent_data = dict(zip(columns, row))

            try:
                # Validate student
                student_admission = parent_data.get("student_admission_number")
                if student_admission:
                    try:
                        student = Student.objects.get(admission_number=student_admission)
                    except Student.DoesNotExist:
                        raise ValueError(f"Row {i}: Student with admission number '{student_admission}' not found")
                else:
                    student = None

                # Check if parent already exists
                if Parent.objects.filter(email=parent_data["email"]).exists():
                    raise ValueError(f"Row {i}: Parent with email '{parent_data['email']}' already exists")

                # Check for duplicate NIN
                if parent_data.get("national_id"):
                    if Parent.objects.filter(national_id=parent_data["national_id"]).exists():
                        raise ValueError(f"Row {i}: Parent with NIN '{parent_data['national_id']}' already exists")

                parent = Parent(
                    first_name=parent_data["first_name"],
                    last_name=parent_data["last_name"],
                    email=parent_data["email"],
                    phone_number=parent_data.get("phone_number"),
                    address=parent_data.get("address"),
                    national_id=parent_data.get("national_id"),
                    relationship=parent_data.get("relationship"),
                    occupation=parent_data.get("occupation"),
                    student=student
                )

                parents_to_create.append(parent)

            except Exception as e:
                results['failed'] += 1
                results['errors'].append(str(e))

        # Bulk create
        with transaction.atomic():
            for parent in parents_to_create:
                parent.save()
                results['created'] += 1

        return results

    except Exception as e:
        results['errors'].append(f"Critical error: {str(e)}")
        return results


@shared_task(name='users.send_email_async')
def send_email_async(subject, message, recipient_list, from_email=None):
    """
    Async task for sending emails.

    Args:
        subject: Email subject
        message: Email body
        recipient_list: List of recipient emails
        from_email: Sender email (optional)

    Returns:
        dict: Send status
    """
    try:
        if from_email is None:
            from_email = settings.DEFAULT_FROM_EMAIL

        send_mail(
            subject=subject,
            message=message,
            from_email=from_email,
            recipient_list=recipient_list,
            fail_silently=False,
        )

        return {
            'status': 'success',
            'message': f'Email sent to {len(recipient_list)} recipient(s)'
        }

    except Exception as e:
        return {
            'status': 'failed',
            'error': str(e)
        }
