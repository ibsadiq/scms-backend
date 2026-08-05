import openpyxl
from django.http import HttpResponse
from django.utils.dateparse import parse_date
from rest_framework import generics, viewsets, status, permissions, serializers
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from rest_framework.response import Response
from .models import AcademicYear, Term, Article, CarouselImage, SchoolEvent
from .serializers import (
    AcademicYearSerializer,
    TermSerializer,
    ArticleSerializer,
    CarouselImageSerializer,
    SchoolEventSerializer,
    SchoolEventBulkUploadSerializer,
)
from .permissions import IsAdminOrReadOnly


from rest_framework.permissions import IsAuthenticated
from django.db.models import Count, Sum, Avg, Q
from django.utils import timezone
from datetime import timedelta
from decimal import Decimal

from academic.models import Subject, Student, Teacher
from finance.models import StudentFeeAssignment, Receipt, FeePaymentAllocation
from attendance.models import StudentAttendance
from examination.models import TermResult


# Article Views
class ArticleListCreateView(generics.ListCreateAPIView):
    queryset = Article.objects.all()
    serializer_class = ArticleSerializer
    permission_classes = [IsAuthenticated]


class ArticleDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Article.objects.all()
    serializer_class = ArticleSerializer
    permission_classes = [IsAuthenticated]


# CarouselImage Views
class CarouselImageListCreateView(generics.ListCreateAPIView):
    queryset = CarouselImage.objects.all()
    serializer_class = CarouselImageSerializer
    permission_classes = [IsAuthenticated]


class CarouselImageDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = CarouselImage.objects.all()
    serializer_class = CarouselImageSerializer
    permission_classes = [IsAuthenticated]


# AcademicYear Views
class AcademicYearListCreateView(generics.ListCreateAPIView):
    queryset = AcademicYear.objects.all()
    serializer_class = AcademicYearSerializer
    permission_classes = [IsAuthenticated]

    def create(self, request, *args, **kwargs):
        print(request.data)
        return super().create(request, *args, **kwargs)


class AcademicYearDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = AcademicYear.objects.all()
    serializer_class = AcademicYearSerializer
    permission_classes = [IsAuthenticated]


# Term Views
class TermListCreateView(generics.ListCreateAPIView):
    serializer_class = TermSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = Term.objects.all()
        academic_year = self.request.query_params.get("academic_year")
        if academic_year:
            queryset = queryset.filter(academic_year_id=academic_year)
        return queryset

    def create(self, request, *args, **kwargs):
        print(request.data)
        return super().create(request, *args, **kwargs)


class TermDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Term.objects.all()
    serializer_class = TermSerializer
    permission_classes = [IsAuthenticated]


from django.utils.dateparse import parse_date
from rest_framework import viewsets

class SchoolEventViewSet(viewsets.ModelViewSet):
    queryset = SchoolEvent.objects.select_related("term", "term__academic_year").all()
    serializer_class = SchoolEventSerializer
    permission_classes = [IsAdminOrReadOnly]

    def get_queryset(self):
        queryset = self.queryset
        term_id = self.request.query_params.get("term")
        year_name = self.request.query_params.get("academic_year")
        event_type = self.request.query_params.get("event_type")
        start_date = parse_date(self.request.query_params.get("start_date") or "")
        end_date = parse_date(self.request.query_params.get("end_date") or "")

        if term_id:
            queryset = queryset.filter(term__id=term_id)
        if year_name:
            queryset = queryset.filter(term__academic_year__name=year_name)
        if event_type:
            queryset = queryset.filter(event_type=event_type)
        
        # FIXED: Overlapping date range logic
        if start_date and end_date:
            # Holiday overlaps range if holiday.start <= range.end AND holiday.end >= range.start
            queryset = queryset.filter(start_date__lte=end_date, end_date__gte=start_date)
        elif start_date:
            queryset = queryset.filter(end_date__gte=start_date)
        elif end_date:
            queryset = queryset.filter(start_date__lte=end_date)

        return queryset

class SchoolEventBulkUploadView(APIView):
    permission_classes = [permissions.IsAdminUser]
    serializer_class = SchoolEventBulkUploadSerializer

    def post(self, request):
        excel_file = request.FILES.get("file")
        if not excel_file:
            return Response(
                {"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST
            )

        try:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
            rows = list(sheet.iter_rows(min_row=2, values_only=True))  # Skip header row

            for row in rows:
                name, event_type, term_id, start_date, end_date, description = row
                if not all([name, event_type, term_id, start_date]):
                    continue  # Skip invalid rows

                SchoolEvent.objects.create(
                    name=name,
                    event_type=event_type,
                    term=Term.objects.get(pk=term_id),
                    start_date=start_date,
                    end_date=end_date,
                    description=description or "",
                )

            return Response(
                {"detail": f"{len(rows)} events uploaded successfully."},
                status=status.HTTP_201_CREATED,
            )

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
class EmptySerializer(serializers.Serializer):
    """Empty serializer for download views that don't require input data"""
    pass

class SchoolEventTemplateDownloadView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = EmptySerializer

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "School Events Template"

        # Headers
        headers = [
            "name",
            "event_type",
            "term_id",
            "start_date",
            "end_date",
            "description",
        ]
        ws.append(headers)

        # Example row
        ws.append(
            [
                "Midterm Exams",
                "exam",
                1,
                "2025-07-10",
                "2025-07-14",
                "Midterm assessment",
            ]
        )

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            "attachment; filename=school_events_template.xlsx"
        )
        wb.save(response)
        return response


class DashboardStatsView(APIView):
    """
    Comprehensive, high-performance admin dashboard stats endpoint.
    GET /api/administration/dashboard/
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from django.core.cache import cache
        from django.db import connection
        from django.db.models import Count, Sum
        from django.db.models.functions import ExtractMonth, ExtractYear

        cache_key = f"admin_dashboard_summary_{connection.schema_name}"
        cached_data = cache.get(cache_key)
        if cached_data:
            return Response(cached_data)

        today = timezone.now().date()
        first_day_of_month = today.replace(day=1)
        
        # ===== BASIC STATS =====
        total_students = Student.objects.filter(is_active=True).count()
        new_students_this_month = Student.objects.filter(
            admission_date__gte=first_day_of_month,
            is_active=True
        ).count()
        total_teachers = Teacher.objects.count()
        active_subjects = Subject.objects.count()
        
        # ===== ATTENDANCE RATE (TODAY) =====
        today_attendance = StudentAttendance.objects.filter(date=today)
        total_expected = Student.objects.filter(is_active=True).count()
        present_count = today_attendance.filter(status__absent=False).count()
        absent_count = today_attendance.filter(status__absent=True).count()
        attendance_rate = round(
            (present_count / total_expected * 100) if total_expected > 0 else 0,
            1
        )

        # ===== STUDENTS BY LEVEL =====
        students_by_level = []
        students_with_class = Student.objects.filter(is_active=True).select_related('class_level')

        class_counts = {}
        for student in students_with_class:
            if hasattr(student, 'class_level') and student.class_level:
                class_name = getattr(student.class_level, 'name', str(student.class_level))
            else:
                class_name = 'Unassigned'
            class_counts[class_name] = class_counts.get(class_name, 0) + 1
        
        # Categorize into levels
        primary_keywords = ['Primary', 'primary', 'P1', 'P2', 'P3', 'P4', 'P5', 'P6']
        jss_keywords = ['JSS', 'jss', 'Junior']
        sss_keywords = ['SSS', 'sss', 'Senior', 'ss', 'SS']
        university_keywords = ['Year', 'year', 'University', 'university']
        
        primary_count = sum(count for class_name, count in class_counts.items() 
                           if any(keyword in class_name for keyword in primary_keywords))
        jss_count = sum(count for class_name, count in class_counts.items() 
                       if any(keyword in class_name for keyword in jss_keywords))
        sss_count = sum(count for class_name, count in class_counts.items() 
                       if any(keyword in class_name for keyword in sss_keywords))
        university_count = sum(count for class_name, count in class_counts.items() 
                              if any(keyword in class_name for keyword in university_keywords))
        
        students_by_level = [
            {
                'name': 'Primary',
                'count': primary_count,
                'percentage': round((primary_count / total_students * 100) if total_students > 0 else 0, 1),
                'icon': 'lucide:baby'
            },
            {
                'name': 'JSS',
                'count': jss_count,
                'percentage': round((jss_count / total_students * 100) if total_students > 0 else 0, 1),
                'icon': 'lucide:book'
            },
            {
                'name': 'SSS',
                'count': sss_count,
                'percentage': round((sss_count / total_students * 100) if total_students > 0 else 0, 1),
                'icon': 'lucide:graduation-cap'
            },
            {
                'name': 'University',
                'count': university_count,
                'percentage': round((university_count / total_students * 100) if total_students > 0 else 0, 1),
                'icon': 'lucide:school'
            }
        ]
        
        # ===== FINANCIAL STATS =====
        try:
            fee_assignments = StudentFeeAssignment.objects.filter(is_waived=False)

            total_expected = fee_assignments.aggregate(
                total=Sum('amount_owed')
            )['total'] or Decimal('0')

            total_paid = fee_assignments.aggregate(
                total=Sum('amount_paid')
            )['total'] or Decimal('0')

            total_outstanding = total_expected - total_paid

            collection_rate = round(
                (float(total_paid) / float(total_expected) * 100)
                if total_expected > 0 else 0, 1
            )

            from django.db.models import F
            students_with_debt = fee_assignments.annotate(
                balance=F('amount_owed') - F('amount_paid')
            ).filter(
                balance__gt=0
            ).values('student').distinct().count()

            financial_stats = {
                'collected': float(total_paid),
                'outstanding': float(total_outstanding),
                'expected': float(total_expected),
                'collectionRate': collection_rate,
                'studentsWithDebt': students_with_debt,
                'totalStudents': total_students
            }
        except Exception as e:
            print(f"Financial calculation error: {e}")
            financial_stats = {
                'collected': 0,
                'outstanding': 0,
                'expected': 0,
                'collectionRate': 0,
                'studentsWithDebt': 0,
                'totalStudents': total_students
            }

        # ===== REVENUE SERIES (LAST 6 MONTHS) =====
        series_months = []
        for i in range(5, -1, -1):
            m = (today.month - i - 1) % 12 + 1
            y = today.year + ((today.month - i - 1) // 12)
            series_months.append((y, m))

        revenue_dict = {}
        try:
            receipts_grouped = Receipt.objects.annotate(
                yr=ExtractYear('date'),
                mo=ExtractMonth('date')
            ).values('yr', 'mo').annotate(total=Sum('amount'))

            for r in receipts_grouped:
                if r['yr'] and r['mo']:
                    revenue_dict[(r['yr'], r['mo'])] = float(r['total'] or 0)
        except Exception:
            pass

        revenue_series = [revenue_dict.get((y, m), 0.0) for y, m in series_months]

        # ===== ENROLLMENT TRENDS (12 Months of Academic Session) =====
        academic_session_months = [
            (9, 2025, 'Sep 2025', 'Sep'),
            (10, 2025, 'Oct 2025', 'Oct'),
            (11, 2025, 'Nov 2025', 'Nov'),
            (12, 2025, 'Dec 2025', 'Dec'),
            (1, 2026, 'Jan 2026', 'Jan'),
            (2, 2026, 'Feb 2026', 'Feb'),
            (3, 2026, 'Mar 2026', 'Mar'),
            (4, 2026, 'Apr 2026', 'Apr'),
            (5, 2026, 'May 2026', 'May'),
            (6, 2026, 'Jun 2026', 'Jun'),
            (7, 2026, 'Jul 2026', 'Jul'),
            (8, 2026, 'Aug 2026', 'Aug'),
        ]
        
        enrollment_dict = {}
        try:
            student_counts = Student.objects.filter(is_active=True, admission_date__isnull=False).annotate(
                yr=ExtractYear('admission_date'),
                mo=ExtractMonth('admission_date')
            ).values('yr', 'mo').annotate(cnt=Count('id'))

            for sc in student_counts:
                if sc['yr'] and sc['mo']:
                    enrollment_dict[(sc['yr'], sc['mo'])] = sc['cnt']
        except Exception:
            pass

        enrollment_trends = [
            {
                'month': m,
                'year': y,
                'label': lbl,
                'shortLabel': slbl,
                'count': enrollment_dict.get((y, m), 0)
            }
            for m, y, lbl, slbl in academic_session_months
        ]

        # ===== RECENT STUDENTS (TOP 10) =====
        recent_students_qs = Student.objects.filter(is_active=True).select_related(
            'classroom', 'classroom__name'
        ).order_by('-admission_date', '-id')[:10]

        recent_students_list = [
            {
                'id': s.id,
                'first_name': s.first_name,
                'last_name': s.last_name,
                'admission_number': s.admission_number or 'N/A',
                'class_name': str(s.classroom) if s.classroom else 'Unassigned',
                'admission_date': s.admission_date.isoformat() if s.admission_date else None,
            }
            for s in recent_students_qs
        ]
        
        # ===== ATTENDANCE FOR THE WEEK =====
        week_start = today - timedelta(days=today.weekday())
        attendance_week = []
        
        for i in range(5):  # Monday to Friday
            day = week_start + timedelta(days=i)
            day_attendance = StudentAttendance.objects.filter(date=day)
            present = day_attendance.filter(status__absent=False).count()
            total = Student.objects.filter(is_active=True).count()
            rate = round((present / total * 100) if total > 0 else 0, 1)
            
            attendance_week.append({
                'dayName': day.strftime('%A'),
                'date': day.isoformat(),
                'rate': rate,
                'present': present,
                'total': total
            })

        # ===== RECENT ADMISSIONS =====
        recent_admissions = Student.objects.filter(
            is_active=True
        ).select_related('class_level').order_by('-admission_date')[:5]

        recent_admissions_list = []
        for student in recent_admissions:
            if hasattr(student, 'class_level') and student.class_level:
                class_name = getattr(student.class_level, 'name', str(student.class_level))
            else:
                class_name = 'Unassigned'

            recent_admissions_list.append({
                'id': student.id,
                'first_name': student.first_name,
                'last_name': student.last_name,
                'grade_level': class_name,
                'admission_date': student.admission_date.isoformat() if student.admission_date else None
            })
        
        # ===== PERFORMANCE STATS =====
        try:
            current_academic_year = AcademicYear.objects.filter(is_current=True).first()

            if current_academic_year:
                current_term = Term.objects.filter(
                    academic_year=current_academic_year
                ).order_by('-start_date').first()
            else:
                current_term = Term.objects.order_by('-start_date').first()

            if current_term:
                term_results = TermResult.objects.filter(term=current_term)
                total_results = term_results.count()

                if total_results > 0:
                    grade_counts = term_results.values('overall_grade').annotate(count=Count('id'))
                    grade_dist = {'A': 0, 'B': 0, 'C': 0, 'D': 0, 'F': 0}

                    for grade_count in grade_counts:
                        grade = grade_count['overall_grade']
                        count = grade_count['count']
                        if grade in grade_dist:
                            grade_dist[grade] = round((count / total_results) * 100)

                    passing_count = term_results.filter(overall_grade__in=['A', 'B', 'C', 'D']).count()
                    pass_rate = round((passing_count / total_results) * 100)

                    most_common_grade = max(grade_counts, key=lambda x: x['count'], default=None)
                    average_grade = most_common_grade['overall_grade'] if most_common_grade else 'N/A'

                    performance_stats = {
                        'averageGrade': average_grade,
                        'passRate': pass_rate,
                        'grades': {
                            'a': grade_dist['A'],
                            'b': grade_dist['B'],
                            'c': grade_dist['C'],
                            'df': grade_dist['D'] + grade_dist['F']
                        }
                    }
                else:
                    performance_stats = {
                        'averageGrade': 'N/A',
                        'passRate': 0,
                        'grades': {'a': 0, 'b': 0, 'c': 0, 'df': 0}
                    }
            else:
                performance_stats = {
                    'averageGrade': 'N/A',
                    'passRate': 0,
                    'grades': {'a': 0, 'b': 0, 'c': 0, 'df': 0}
                }
        except Exception as e:
            performance_stats = {
                'averageGrade': 'N/A',
                'passRate': 0,
                'grades': {'a': 0, 'b': 0, 'c': 0, 'df': 0}
            }
        
        # ===== RECENT PAYMENTS =====
        try:
            recent_receipts = Receipt.objects.select_related(
                'student', 'term'
            ).order_by('-date')[:5]

            recent_payments_list = []
            for receipt in recent_receipts:
                recent_payments_list.append({
                    'id': receipt.id,
                    'receipt_number': receipt.receipt_number,
                    'student_name': receipt.student.full_name if receipt.student else receipt.payer,
                    'amount': float(receipt.amount),
                    'method': receipt.paid_through,
                    'paid_on': receipt.payment_date.isoformat() if receipt.payment_date else receipt.date.isoformat(),
                    'term_name': receipt.term.name if receipt.term else 'N/A'
                })
        except Exception as e:
            print(f"Recent payments error: {e}")
            recent_payments_list = []
        
        # ===== COMPILE RESPONSE PAYLOAD =====
        response_payload = {
            'stats': {
                'totalStudents': total_students,
                'totalTeachers': total_teachers,
                'activeSubjects': active_subjects,
                'attendanceRate': attendance_rate,
                'attendancePresent': present_count,
                'attendanceAbsent': absent_count,
                'newStudentsThisMonth': new_students_this_month,
                'revenueCollected': financial_stats['collected'],
                'pendingFees': financial_stats['outstanding'],
                'revenueSeries': revenue_series,
            },
            'enrollmentTrends': enrollment_trends,
            'recentStudents': recent_students_list,
            'studentsByLevel': students_by_level,
            'financial': financial_stats,
            'attendance': attendance_week,
            'recentAdmissions': recent_admissions_list,
            'recentPayments': recent_payments_list,
            'performance': performance_stats
        }

        # Cache payload for 30 seconds
        cache.set(cache_key, response_payload, 30)

        return Response(response_payload)
