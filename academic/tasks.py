"""
Celery tasks for academic operations.

Handles asynchronous operations like:
- Bulk student uploads
- Bulk classroom creation
- Class advancement processing
"""
from celery import shared_task
from django.core.exceptions import ValidationError
from django.db import transaction
import openpyxl
from io import BytesIO

from academic.models import (
    Student, ClassRoom, GradeLevel, Stream, Teacher,
    StudentClassEnrollment
)
from users.models import CustomUser as User
from administration.models import AcademicYear
from django_tenants.utils import schema_context


@shared_task(bind=True, name='academic.bulk_upload_students')
def bulk_upload_students_task(self, schema_name, file_content, academic_year_id, uploaded_by_id):
    """
    Async task for bulk uploading students from Excel file.

    Args:
        self: Celery task instance
        schema_name: The tenant schema to execute in
        file_content: Binary file content
        academic_year_id: ID of academic year for enrollment
        uploaded_by_id: ID of user who initiated upload

    Returns:
        dict: Summary of upload results
    """
    with schema_context(schema_name):
        results = {
            'total_rows': 0,
            'created': 0,
            'failed': 0,
            'errors': []
        }

        try:
            # Load workbook from bytes
            workbook = openpyxl.load_workbook(BytesIO(file_content))
            sheet = workbook.active

            columns = [
                "admission_number",
                "first_name",
                "last_name",
                "date_of_birth",
                "gender",
                "email",
                "phone_number",
                "address",
                "classroom_name",
                "medical_conditions",
            ]

            students_to_create = []
            enrollments_to_create = []

            for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                results['total_rows'] += 1

                # Update task progress
                self.update_state(
                    state='PROGRESS',
                    meta={'current': i - 1, 'total': sheet.max_row - 1, 'status': f'Processing row {i}'}
                )

                student_data = dict(zip(columns, row))

                try:
                    # Validate classroom
                    classroom_name = student_data.get("classroom_name")
                    if not classroom_name:
                        raise ValueError(f"Row {i}: Classroom name is required")

                    classroom = ClassRoom.objects.filter(
                        name__iexact=classroom_name
                    ).first()

                    if not classroom:
                        raise ValueError(f"Row {i}: Classroom '{classroom_name}' does not exist")

                    # Check classroom capacity
                    if classroom.occupied_sits >= classroom.capacity:
                        raise ValueError(f"Row {i}: Classroom '{classroom_name}' has reached its capacity")

                    # Check if student already exists
                    if Student.objects.filter(
                        admission_number=student_data["admission_number"]
                    ).exists():
                        raise ValueError(f"Row {i}: Student with admission number already exists")

                    # Create student instance
                    student = Student(
                        admission_number=student_data["admission_number"],
                        first_name=student_data["first_name"],
                        last_name=student_data["last_name"],
                        date_of_birth=student_data["date_of_birth"],
                        gender=student_data["gender"].upper()[0],
                        email=student_data.get("email"),
                        phone_number=student_data.get("phone_number"),
                        address=student_data.get("address"),
                        medical_conditions=student_data.get("medical_conditions"),
                    )

                    students_to_create.append({
                        'student': student,
                        'classroom': classroom,
                        'row': i
                    })

                except Exception as e:
                    results['failed'] += 1
                    results['errors'].append(str(e))

            # Bulk create students and enrollments
            with transaction.atomic():
                academic_year = AcademicYear.objects.get(id=academic_year_id)

                for item in students_to_create:
                    try:
                        student = item['student']
                        classroom = item['classroom']

                        student.save()

                        # Create enrollment
                        from academic.services.enrollment_service import EnrollmentService
                        EnrollmentService.enroll(
                            student=student, classroom=classroom, academic_year=academic_year
                        )

                        results['created'] += 1

                    except Exception as e:
                        results['failed'] += 1
                        results['errors'].append(f"Row {item['row']}: {str(e)}")

            return results

        except Exception as e:
            results['errors'].append(f"Critical error: {str(e)}")
            return results


@shared_task(bind=True, name='academic.bulk_upload_classrooms')
def bulk_upload_classrooms_task(self, schema_name, file_content):
    """
    Async task for bulk uploading classrooms from Excel file.

    Args:
        self: Celery task instance
        schema_name: The tenant schema to execute in
        file_content: Binary file content

    Returns:
        dict: Summary of upload results
    """
    with schema_context(schema_name):
        results = {
            'total_rows': 0,
            'created': 0,
            'failed': 0,
            'errors': []
        }

        try:
            workbook = openpyxl.load_workbook(BytesIO(file_content))
            sheet = workbook.active

            columns = ["grade_level", "name", "stream", "class_teacher"]

            classrooms_to_create = []

            for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                results['total_rows'] += 1

                # Update progress
                self.update_state(
                    state='PROGRESS',
                    meta={'current': i - 1, 'total': sheet.max_row - 1, 'status': f'Processing row {i}'}
                )

                classroom_data = dict(zip(columns, row))

                try:
                    # Validate GradeLevel
                    gl_raw = str(classroom_data.get("grade_level") or "").strip()
                    grade_level = (
                        GradeLevel.objects.filter(system_code__iexact=gl_raw).first()
                        or GradeLevel.objects.filter(alias__iexact=gl_raw).first()
                        or GradeLevel.objects.filter(default_name__iexact=gl_raw).first()
                    )
                    if not grade_level:
                        raise ValueError(f"Row {i}: Grade level '{gl_raw}' not found")

                    name_str = str(classroom_data.get("name") or "").strip()
                    if not name_str:
                        raise ValueError(f"Row {i}: Classroom name is required")

                    # Handle stream (optional)
                    stream = None
                    if classroom_data.get("stream") and str(classroom_data["stream"]).strip():
                        stream_name = str(classroom_data["stream"]).strip().upper()
                        stream = Stream.objects.filter(name=stream_name).first()
                        if not stream:
                            stream = Stream.objects.create(name=stream_name)

                    # Check if classroom already exists
                    if ClassRoom.objects.filter(grade_level=grade_level, stream=stream, name__iexact=name_str).exists():
                        raise ValueError(f"Row {i}: Classroom '{name_str}' already exists for {grade_level}")

                    # Validate teacher (optional)
                    class_teacher = None
                    if classroom_data.get("class_teacher") and str(classroom_data["class_teacher"]).strip():
                        teacher_name = str(classroom_data["class_teacher"]).strip()
                        if " " in teacher_name:
                            first_name, last_name = teacher_name.rsplit(" ", 1)
                            class_teacher = Teacher.objects.filter(
                                first_name__iexact=first_name.strip(),
                                last_name__iexact=last_name.strip()
                            ).first()

                    classroom = ClassRoom(
                        name=name_str,
                        grade_level=grade_level,
                        stream=stream,
                        class_teacher=class_teacher
                    )

                    classrooms_to_create.append(classroom)

                except Exception as e:
                    results['failed'] += 1
                    results['errors'].append(f"Row {i}: {str(e)}")

            # Bulk create
            with transaction.atomic():
                for classroom in classrooms_to_create:
                    classroom.save()
                    results['created'] += 1

            return results

        except Exception as e:
            results['errors'].append(f"Critical error: {str(e)}")
            return results
