import openpyxl
from django.db.models import F
from django.core.exceptions import ValidationError
from rest_framework import generics, status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.generics import ListCreateAPIView, RetrieveUpdateDestroyAPIView
from academic.permissions import IsAcademicAdminOrReadOnly, IsSchoolAdmin

from academic.models import (
    Department,
    GradeLevel,
    SchoolSection,
    ClassYear,
    ReasonLeft,
    Stream,
    ClassRoom,
    Subject,
    Teacher,
)
from academic.serializers import (
    DepartmentSerializer,
    GradeLevelSerializer,
    SchoolSectionSerializer,
    ClassYearSerializer,
    ReasonLeftSerializer,
    StreamSerializer,
    ClassRoomSerializer,
    BulkUploadClassRoomsSerializer,
    BulkUploadSubjectsSerializer,
)


class DepartmentListCreateView(generics.ListCreateAPIView):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer
    permission_classes = [IsAcademicAdminOrReadOnly]


class DepartmentDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer
    permission_classes = [IsAcademicAdminOrReadOnly]


class SchoolSectionListView(generics.ListAPIView):
    queryset = SchoolSection.objects.all()
    serializer_class = SchoolSectionSerializer
    permission_classes = [IsAcademicAdminOrReadOnly]


class SchoolSectionDetailView(generics.RetrieveUpdateAPIView):
    queryset = SchoolSection.objects.all()
    serializer_class = SchoolSectionSerializer
    permission_classes = [IsAcademicAdminOrReadOnly]


class GradeLevelListCreateView(generics.ListCreateAPIView):
    queryset = GradeLevel.objects.all()
    serializer_class = GradeLevelSerializer
    permission_classes = [IsAcademicAdminOrReadOnly]


class GradeLevelDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = GradeLevel.objects.all()
    serializer_class = GradeLevelSerializer
    permission_classes = [IsAcademicAdminOrReadOnly]


class ClassYearListCreateView(generics.ListCreateAPIView):
    queryset = ClassYear.objects.all()
    serializer_class = ClassYearSerializer
    permission_classes = [IsAcademicAdminOrReadOnly]


class ClassYearDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = ClassYear.objects.all()
    serializer_class = ClassYearSerializer
    permission_classes = [IsAcademicAdminOrReadOnly]


class ReasonLeftListCreateView(generics.ListCreateAPIView):
    queryset = ReasonLeft.objects.all()
    serializer_class = ReasonLeftSerializer
    permission_classes = [IsAcademicAdminOrReadOnly]


class ReasonLeftDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = ReasonLeft.objects.all()
    serializer_class = ReasonLeftSerializer
    permission_classes = [IsAcademicAdminOrReadOnly]


class StreamListCreateView(generics.ListCreateAPIView):
    queryset = Stream.objects.all()
    serializer_class = StreamSerializer
    permission_classes = [IsAcademicAdminOrReadOnly]


class StreamDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Stream.objects.all()
    serializer_class = StreamSerializer
    permission_classes = [IsAcademicAdminOrReadOnly]


class ClassRoomView(ListCreateAPIView):
    serializer_class = ClassRoomSerializer
    queryset = ClassRoom.objects.all()
    permission_classes = [IsAcademicAdminOrReadOnly]


class ClassRoomDetailView(RetrieveUpdateDestroyAPIView):
    serializer_class = ClassRoomSerializer
    queryset = ClassRoom.objects.all()
    permission_classes = [IsAcademicAdminOrReadOnly]


class BulkUploadClassRoomsView(APIView):
    serializer_class = BulkUploadClassRoomsSerializer
    permission_classes = [IsSchoolAdmin]

    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            return Response(
                {"error": "No file provided."}, status=status.HTTP_400_BAD_REQUEST
            )

        try:
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            columns = [
                "grade_level",
                "name",
                "stream",
                "class_teacher",
            ]

            classrooms_to_create = []
            not_created = []

            for i, row in enumerate(
                sheet.iter_rows(min_row=2, values_only=True), start=2
            ):
                classroom_data = dict(zip(columns, row))

                try:
                    gl_raw = str(classroom_data.get("grade_level") or "").strip()
                    grade_level = (
                        GradeLevel.objects.filter(system_code__iexact=gl_raw).first()
                        or GradeLevel.objects.filter(alias__iexact=gl_raw).first()
                        or GradeLevel.objects.filter(default_name__iexact=gl_raw).first()
                    )
                    if not grade_level:
                        raise ValueError(f"Row {i}: Grade level '{gl_raw}' not found.")

                    name_str = str(classroom_data.get("name") or "").strip()
                    if not name_str:
                        raise ValueError(f"Row {i}: Classroom name is required.")

                    stream = None
                    if classroom_data.get("stream") and str(classroom_data["stream"]).strip():
                        stream_raw = str(classroom_data["stream"]).strip().upper()
                        try:
                            stream = Stream.objects.get(name=stream_raw)
                        except Stream.DoesNotExist:
                            stream = Stream.objects.create(name=stream_raw)

                    if ClassRoom.objects.filter(grade_level=grade_level, stream=stream, name__iexact=name_str).exists():
                        raise ValueError(
                            f"Row {i}: ClassRoom '{name_str}' already exists for grade '{grade_level}'."
                        )

                    class_teacher = None
                    if classroom_data.get("class_teacher") and str(classroom_data["class_teacher"]).strip():
                        teacher_name_parts = str(classroom_data["class_teacher"]).lower().split()
                        if len(teacher_name_parts) >= 2:
                            class_teacher_first_name = teacher_name_parts[0]
                            class_teacher_last_name = teacher_name_parts[1]
                            class_teacher = Teacher.objects.filter(
                                first_name__iexact=class_teacher_first_name,
                                last_name__iexact=class_teacher_last_name,
                            ).first()

                    classroom = ClassRoom(
                        name=name_str,
                        grade_level=grade_level,
                        stream=stream,
                        class_teacher=class_teacher,
                    )
                    classrooms_to_create.append(classroom)

                except Exception as e:
                    not_created.append(
                        {
                            "row": i,
                            "data": classroom_data,
                            "error": str(e),
                        }
                    )

            if classrooms_to_create:
                ClassRoom.objects.bulk_create(classrooms_to_create)

            return Response(
                {
                    "created": len(classrooms_to_create),
                    "not_created": not_created,
                },
                status=status.HTTP_201_CREATED,
            )

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class BulkUploadSubjectsView(APIView):
    permission_classes = [IsSchoolAdmin]
    serializer_class = BulkUploadSubjectsSerializer

    def post(self, request, *args, **kwargs):
        file = request.FILES.get("file")
        if not file:
            return Response(
                {"error": "No file provided."}, status=status.HTTP_400_BAD_REQUEST
            )

        try:
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            columns = [
                "name",
                "subject_code",
                "department",
            ]

            subjects_to_create = []
            not_created = []

            for i, row in enumerate(
                sheet.iter_rows(min_row=2, values_only=True), start=2
            ):
                subject_data = dict(zip(columns, row))

                try:
                    try:
                        department = Department.objects.get(
                            name=subject_data["department"].strip().lower()
                        )
                    except Department.DoesNotExist:
                        raise ValueError(
                            f"Department '{subject_data['department']}' does not exist."
                        )

                    if Subject.objects.filter(
                        subject_code=subject_data["subject_code"]
                    ).exists():
                        raise ValueError(
                            f"Subject code '{subject_data['subject_code']}' already exists."
                        )
                    if Subject.objects.filter(name=subject_data["name"]).exists():
                        raise ValueError(
                            f"Subject name '{subject_data['name']}' already exists."
                        )

                    subject = Subject(
                        name=subject_data["name"].strip(),
                        subject_code=subject_data["subject_code"].strip(),
                        description=f"{subject_data['name']} ({subject_data['subject_code']})",
                        department=department,
                    )
                    subjects_to_create.append(subject)

                except Exception as e:
                    subject_data["error"] = str(e)
                    not_created.append(subject_data)

            if subjects_to_create:
                Subject.objects.bulk_create(subjects_to_create)

            return Response(
                {
                    "message": f"{len(subjects_to_create)} subjects successfully uploaded.",
                    "not_created": not_created,
                },
                status=status.HTTP_201_CREATED,
            )

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
