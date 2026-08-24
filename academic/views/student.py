import openpyxl
from django.core.exceptions import ValidationError
from rest_framework import generics, status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from academic.permissions import IsAcademicAdminOrReadOnly, IsSchoolAdmin

from administration.models import AcademicYear
from academic.models import (
    StudentClassEnrollment as StudentClass,
    Student,
    ClassRoom,
    ClassLevel,
    Parent,
)
from academic.serializers import (
    StudentClassEnrollmentSerializer,
    BulkUploadStudentsSerializer,
    BulkCreateStudentsProfileSerializer,
)


class StudentClassListCreateView(generics.ListCreateAPIView):
    queryset = StudentClass.objects.all()
    serializer_class = StudentClassEnrollmentSerializer
    permission_classes = [IsAcademicAdminOrReadOnly]


class StudentClassDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = StudentClass.objects.all()
    serializer_class = StudentClassEnrollmentSerializer
    permission_classes = [IsAcademicAdminOrReadOnly]


class BulkUploadStudentClassView(APIView):
    serializer_class = BulkUploadStudentsSerializer
    permission_classes = [IsSchoolAdmin]

    def post(self, request, *args, **kwargs):
        file = request.FILES.get("file")
        if not file:
            return Response(
                {"error": "No file provided."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            columns = [
                "classroom_name",
                "academic_year",
                "student_full_name",
            ]

            rows_with_errors = []
            student_classes_to_create = []
            for i, row in enumerate(
                sheet.iter_rows(min_row=2, values_only=True), start=2
            ):
                row_data = dict(zip(columns, row))

                try:
                    try:
                        classroom = ClassRoom.objects.get(
                            name__name=row_data["classroom_name"].strip().lower(),
                        )
                    except ClassRoom.DoesNotExist:
                        raise ValidationError(
                            f"Row {i}: Classroom '{row_data['classroom_name']}' does not exist."
                        )

                    try:
                        academic_year = AcademicYear.objects.get(
                            name=row_data["academic_year"]
                        )
                    except AcademicYear.DoesNotExist:
                        raise ValidationError(
                            f"Row {i}: Academic year '{row_data['academic_year']}' does not exist."
                        )

                    full_name = row_data["student_full_name"].strip().lower()
                    name_parts = full_name.split()
                    if len(name_parts) < 2:
                        raise ValidationError(
                            f"Row {i}: Student full name '{full_name}' must have at least a first name and last name."
                        )

                    first_name, last_name = name_parts[0], name_parts[-1]

                    try:
                        student = Student.objects.get(
                            first_name__iexact=first_name,
                            last_name__iexact=last_name,
                        )
                    except Student.DoesNotExist:
                        raise ValidationError(
                            f"Row {i}: No student found with the name '{full_name}'."
                        )

                    if classroom.name != student.class_level:
                        raise ValidationError(
                            f"Row {i}: The classroom '{row_data['classroom_name']}' does not match the student's class level '{student.class_level}'."
                        )

                    if StudentClass.objects.filter(
                        classroom=classroom,
                        academic_year=academic_year,
                        student=student,
                    ).exists():
                        raise ValidationError(
                            f"Row {i}: Student '{full_name}' is already assigned to this class in the given academic year."
                        )

                    if classroom.occupied_sits >= classroom.capacity:
                        raise ValidationError(
                            f"Row {i}: Classroom '{classroom}' has reached its capacity."
                        )

                    student_class = StudentClass(
                        classroom=classroom,
                        academic_year=academic_year,
                        student=student,
                    )
                    student_classes_to_create.append(student_class)

                except ValidationError as e:
                    rows_with_errors.append({"row": i, "errors": str(e)})

            from academic.services.enrollment_service import EnrollmentService
            EnrollmentService.bulk_enroll([
                {
                    "student": student_class.student,
                    "classroom": student_class.classroom,
                    "academic_year": student_class.academic_year,
                }
                for student_class in student_classes_to_create
            ])

            return Response(
                {
                    "message": f"{len(student_classes_to_create)} student-class records successfully uploaded.",
                    "errors": rows_with_errors,
                },
                status=status.HTTP_201_CREATED,
            )

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class BulkUploadStudentsProfileView(APIView):
    serializer_class = BulkCreateStudentsProfileSerializer
    permission_classes = [IsSchoolAdmin]

    def post(self, request, *args, **kwargs):
        file = request.FILES.get("file")
        if not file:
            return Response({"error": "No file provided."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            columns = [
                "first_name",
                "last_name",
                "gender",
                "date_of_birth",
                "class_level",
                "parent_name",
                "parent_contact",
            ]

            rows_with_errors = []
            created_count = 0

            for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                row_data = dict(zip(columns, row))
                if not any(row_data.values()):
                    continue

                try:
                    if not row_data.get("first_name") or not row_data.get("last_name"):
                        raise ValidationError(f"Row {i}: First Name and Last Name are required.")
                    if not row_data.get("class_level"):
                        raise ValidationError(f"Row {i}: Class Level is required.")

                    try:
                        class_level = ClassLevel.objects.get(name__iexact=str(row_data["class_level"]).strip())
                    except ClassLevel.DoesNotExist:
                        raise ValidationError(f"Row {i}: Class Level '{row_data['class_level']}' does not exist.")

                    parent = None
                    if row_data.get("parent_contact"):
                        contact = str(row_data["parent_contact"]).strip()
                        parent_name = str(row_data.get("parent_name") or "").strip()
                        parent = Parent.objects.filter(children__parent_contact=contact).first()
                        if not parent:
                            parent = Parent.objects.create(first_name=parent_name)

                    student = Student(
                        first_name=str(row_data["first_name"]).strip(),
                        last_name=str(row_data["last_name"]).strip(),
                        gender=str(row_data.get("gender") or "").strip().capitalize(),
                        date_of_birth=row_data.get("date_of_birth"),
                        class_level=class_level,
                        parent_guardian=parent,
                        parent_contact=str(row_data.get("parent_contact") or "").strip(),
                        is_active=True,
                    )
                    student.save()
                    created_count += 1
                except ValidationError as e:
                    rows_with_errors.append({"row": i, "errors": str(e)})
                except Exception as e:
                    rows_with_errors.append({"row": i, "errors": str(e)})

            return Response(
                {
                    "message": f"{created_count} students successfully created.",
                    "errors": rows_with_errors,
                },
                status=status.HTTP_201_CREATED,
            )

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
