import openpyxl
from django.db.models import F
from django.core.exceptions import ValidationError
from rest_framework import generics
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.generics import ListCreateAPIView, RetrieveUpdateDestroyAPIView
from administration.models import AcademicYear
from .models import (
    Subject,
    Department,
    ClassRoom,
    ClassLevel,
    Teacher,
    GradeLevel,
    ClassYear,
    ReasonLeft,
    StudentClassEnrollment as StudentClass,
    Student,
    Stream,
)
from .serializers import (
    DepartmentSerializer,
    ClassLevelSerializer,
    GradeLevelSerializer,
    ClassYearSerializer,
    ReasonLeftSerializer,
    SubjectSerializer,
    ClassRoomSerializer,
    StudentClassEnrollmentSerializer,
    BulkUploadClassRoomsSerializer,
    BulkUploadStudentsSerializer,
    BulkUploadSubjectsSerializer,
    StreamSerializer,
)


# Department Views
class DepartmentListCreateView(generics.ListCreateAPIView):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer
    permission_classes = [IsAuthenticated]


class DepartmentDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer
    permission_classes = [IsAuthenticated]


# ClassLevel Views
class ClassLevelListCreateView(generics.ListCreateAPIView):
    queryset = ClassLevel.objects.all()
    serializer_class = ClassLevelSerializer
    permission_classes = [IsAuthenticated]

    def create(self, request, *args, **kwargs):
        print(request.data)
        return super().create(request, *args, **kwargs)


class ClassLevelDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = ClassLevel.objects.all()
    serializer_class = ClassLevelSerializer
    permission_classes = [IsAuthenticated]


# GradeLevel Views
class GradeLevelListCreateView(generics.ListCreateAPIView):
    queryset = GradeLevel.objects.all()
    serializer_class = GradeLevelSerializer
    permission_classes = [IsAuthenticated]


class GradeLevelDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = GradeLevel.objects.all()
    serializer_class = GradeLevelSerializer
    permission_classes = [IsAuthenticated]


# ClassYear Views
class ClassYearListCreateView(generics.ListCreateAPIView):
    queryset = ClassYear.objects.all()
    serializer_class = ClassYearSerializer
    permission_classes = [IsAuthenticated]


class ClassYearDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = ClassYear.objects.all()
    serializer_class = ClassYearSerializer
    permission_classes = [IsAuthenticated]


# ReasonLeft Views
class ReasonLeftListCreateView(generics.ListCreateAPIView):
    queryset = ReasonLeft.objects.all()
    serializer_class = ReasonLeftSerializer
    permission_classes = [IsAuthenticated]


class ReasonLeftDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = ReasonLeft.objects.all()
    serializer_class = ReasonLeftSerializer
    permission_classes = [IsAuthenticated]


# Stream Views
class StreamListCreateView(generics.ListCreateAPIView):
    queryset = Stream.objects.all()
    serializer_class = StreamSerializer
    permission_classes = [IsAuthenticated]


class StreamDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Stream.objects.all()
    serializer_class = StreamSerializer
    permission_classes = [IsAuthenticated]


class SubjectListView(generics.ListCreateAPIView):
    queryset = Subject.objects.all()
    serializer_class = SubjectSerializer
    permission_classes = [IsAuthenticated]


class SubjectDetailView(generics.RetrieveUpdateDestroyAPIView):
    """
    API view to retrieve, update, or delete a subject.
    """

    queryset = Subject.objects.all()
    serializer_class = SubjectSerializer
    lookup_field = "id"  # You can change this to "subject_code" if needed


class BulkUploadSubjectsView(APIView):
    """
    API View to handle bulk uploading of subjects from an Excel file.
    """
    serializer_class = BulkUploadSubjectsSerializer

    def post(self, request, *args, **kwargs):
        file = request.FILES.get("file")
        if not file:
            return Response(
                {"error": "No file provided."}, status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Load the Excel file
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active  # Assuming data is in the first sheet

            # Expected columns in the Excel file
            columns = [
                "name",  # Subject name
                "subject_code",  # Subject code
                "department",  # Department name (for ForeignKey)
            ]

            subjects_to_create = []
            not_created = []  # List to store invalid rows and their errors

            for i, row in enumerate(
                sheet.iter_rows(min_row=2, values_only=True), start=2
            ):
                # Map row data to the expected columns
                subject_data = dict(zip(columns, row))

                try:
                    # Validate department
                    try:
                        department = Department.objects.get(
                            name=subject_data["department"].strip().lower()
                        )
                    except Department.DoesNotExist:
                        raise ValueError(
                            f"Department '{subject_data['department']}' does not exist."
                        )

                    # Check for duplicate subject based on subject code or name
                    if Subject.objects.filter(
                        subject_code=subject_data["subject_code"]
                    ).exists():
                        raise ValueError(
                            f"Subject code '{subject_data['subject_code']}' already exists."
                        )
                    if Subject.objects.filter(name=subject_data["name"]).exists():
                        raise ValueError(
                            f"Subject name '{subject_data['name']}' already exists."
                        )

                    # Create Subject object
                    subject = Subject(
                        name=subject_data["name"].strip(),
                        subject_code=subject_data["subject_code"].strip(),
                        description=f"{subject_data['name']} ({subject_data['subject_code']})",
                        department=department,
                    )
                    subjects_to_create.append(subject)

                except Exception as e:
                    # Add row data and error message to the not_created list
                    subject_data["error"] = str(e)
                    not_created.append(subject_data)

            # Bulk create valid subjects
            if subjects_to_create:
                Subject.objects.bulk_create(subjects_to_create)

            return Response(
                {
                    "message": f"{len(subjects_to_create)} subjects successfully uploaded.",
                    "not_created": not_created,
                },
                status=status.HTTP_201_CREATED,
            )

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


# class ClassRoomView(APIView):
#     """
#     API View to handle CRUD operations for ClassRoom model.
#     """

#     def get(self, request):
#         classrooms = ClassRoom.objects.all()
#         serializer = ClassRoomSerializer(classrooms, many=True)
#         return Response(serializer.data, status=status.HTTP_200_OK)

#     def post(self, request):
#         serializer = ClassRoomSerializer(data=request.data)
#         if serializer.is_valid():
#             serializer.save()
#             return Response(serializer.data, status=status.HTTP_201_CREATED)
#         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class ClassRoomView(ListCreateAPIView):
    """Handle listing and creating classrooms"""
    serializer_class = ClassRoomSerializer
    queryset = ClassRoom.objects.all()

class ClassRoomDetailView(RetrieveUpdateDestroyAPIView):
    """Handle retrieve, update, delete for single classroom"""
    serializer_class = ClassRoomSerializer
    queryset = ClassRoom.objects.all()


class BulkUploadClassRoomsView(APIView):
    """
    API View to handle bulk uploading of classrooms from an Excel file.
    """

    serializer_class = BulkUploadClassRoomsSerializer

    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            return Response(
                {"error": "No file provided."}, status=status.HTTP_400_BAD_REQUEST
            )

        try:
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active  # Assuming data is in the first sheet

            columns = [
                "name",  # ClassLevel name
                "stream",  # Stream name (optional)
                "class_teacher",  # Full name of class teacher
            ]

            classrooms_to_create = []
            not_created = []

            for i, row in enumerate(
                sheet.iter_rows(min_row=2, values_only=True), start=2
            ):
                classroom_data = dict(zip(columns, row))

                try:
                    # Validate and fetch related objects
                    name = ClassLevel.objects.get(name=classroom_data["name"].lower())

                    # Check if ClassRoom with the same name already exists
                    if ClassRoom.objects.filter(name=name).exists():
                        raise ValueError(
                            f"Row {i}: ClassRoom with the same name already exists."
                        )

                    # Handle stream (optional)
                    stream = None
                    if classroom_data.get("stream") and classroom_data["stream"].strip():
                        try:
                            stream = Stream.objects.get(name=classroom_data["stream"].strip().upper())
                        except Stream.DoesNotExist:
                            raise ValueError(
                                f"Stream '{classroom_data['stream']}' does not exist."
                            )

                    # Split the class teacher name and fetch the teacher
                    teacher_name_parts = classroom_data["class_teacher"].lower().split()
                    if len(teacher_name_parts) < 2:
                        raise ValueError(
                            "Class teacher name must include both first and last name."
                        )

                    class_teacher_first_name = teacher_name_parts[0]
                    class_teacher_last_name = teacher_name_parts[1]

                    class_teacher = Teacher.objects.get(
                        first_name=class_teacher_first_name,
                        last_name=class_teacher_last_name,
                    )

                    # Create the ClassRoom object
                    classroom = ClassRoom(
                        name=name,
                        stream=stream,
                        class_teacher=class_teacher,
                    )
                    classrooms_to_create.append(classroom)

                except Exception as e:
                    # Collect the row and its error
                    not_created.append(
                        {
                            "row": i,
                            "data": classroom_data,
                            "error": str(e),
                        }
                    )

            # Bulk create classrooms
            if classrooms_to_create:
                ClassRoom.objects.bulk_create(classrooms_to_create)

            return Response(
                {
                    "created": len(classrooms_to_create),
                    "not_created": not_created,
                },
                status=status.HTTP_201_CREATED,
            )

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class StudentClassListCreateView(generics.ListCreateAPIView):
    """
    View to list and create student-class assignments.
    """

    queryset = StudentClass.objects.all()
    serializer_class = StudentClassEnrollmentSerializer
    permission_classes = [IsAuthenticated]


class StudentClassDetailView(generics.RetrieveUpdateDestroyAPIView):
    """
    View to retrieve, update, or delete a specific student-class assignment.
    """

    queryset = StudentClass.objects.all()
    serializer_class = StudentClassEnrollmentSerializer
    permission_classes = [IsAuthenticated]


class BulkUploadStudentClassView(APIView):
    """
    API View to handle bulk uploading of StudentClass records from an Excel file.
    """
    serializer_class = BulkUploadStudentsSerializer
    
    def post(self, request, *args, **kwargs):
        file = request.FILES.get("file")
        if not file:
            return Response(
                {"error": "No file provided."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            # Load the Excel file
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active  # Assuming the data is in the first sheet

            # Expected columns in the Excel file
            columns = [
                "classroom_name",  # Classroom name
                "academic_year",  # Academic year
                "student_full_name",  # Student's full name
            ]

            rows_with_errors = []
            student_classes_to_create = []
            for i, row in enumerate(
                sheet.iter_rows(min_row=2, values_only=True), start=2
            ):
                # Map row data to the expected columns
                row_data = dict(zip(columns, row))

                try:
                    # Validate Classroom
                    try:
                        classroom = ClassRoom.objects.get(
                            name__name=row_data["classroom_name"].strip().lower(),
                        )
                    except ClassRoom.DoesNotExist:
                        raise ValidationError(
                            f"Row {i}: Classroom '{row_data['classroom_name']}' does not exist."
                        )

                    # Validate AcademicYear
                    try:
                        academic_year = AcademicYear.objects.get(
                            name=row_data["academic_year"]
                        )
                    except AcademicYear.DoesNotExist:
                        raise ValidationError(
                            f"Row {i}: Academic year '{row_data['academic_year']}' does not exist."
                        )

                    # Validate Student by splitting the full name
                    full_name = row_data["student_full_name"].strip().lower()
                    name_parts = full_name.split()
                    if len(name_parts) < 2:
                        raise ValidationError(
                            f"Row {i}: Student full name '{full_name}' must have at least a first name and last name."
                        )

                    first_name, last_name = name_parts[0], name_parts[-1]

                    try:
                        student = Student.objects.get(
                            first_name__iexact=first_name,
                            last_name__iexact=last_name,
                        )
                    except Student.DoesNotExist:
                        raise ValidationError(
                            f"Row {i}: No student found with the name '{full_name}'."
                        )

                    # Validate that the classroom matches the student's class_level
                    if classroom.name != student.class_level:
                        raise ValidationError(
                            f"Row {i}: The classroom '{row_data['classroom_name']}' does not match the student's class level '{student.class_level}'."
                        )

                    # Check for duplicate StudentClass
                    if StudentClass.objects.filter(
                        classroom=classroom,
                        academic_year=academic_year,
                        student=student,
                    ).exists():
                        raise ValidationError(
                            f"Row {i}: Student '{full_name}' is already assigned to this class in the given academic year."
                        )

                    # Check for class capacity
                    if classroom.occupied_sits >= classroom.capacity:
                        raise ValidationError(
                            f"Row {i}: Classroom '{classroom}' has reached its capacity."
                        )

                    # Create StudentClass object
                    student_class = StudentClass(
                        classroom=classroom,
                        academic_year=academic_year,
                        student=student,
                    )
                    student_classes_to_create.append(student_class)

                except ValidationError as e:
                    rows_with_errors.append({"row": i, "errors": str(e)})

            # Bulk create valid records
            StudentClass.objects.bulk_create(student_classes_to_create)

            # Update the occupied_sits atomically for all classrooms
            for student_class in student_classes_to_create:
                student_class.classroom.occupied_sits += 1
                student_class.classroom.save()

            return Response(
                {
                    "message": f"{len(student_classes_to_create)} student-class records successfully uploaded.",
                    "errors": rows_with_errors,
                },
                status=status.HTTP_201_CREATED,
            )

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
