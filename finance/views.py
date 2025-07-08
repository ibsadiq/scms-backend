import openpyxl
from openpyxl.styles import Font
from django.db import transaction
from django.http import Http404, HttpResponse
from django.shortcuts import get_object_or_404
from django_filters.rest_framework import (
    DjangoFilterBackend,
    FilterSet,
    DateFilter,
    CharFilter,
)
from rest_framework import generics, status
from rest_framework.exceptions import NotFound
from rest_framework.filters import SearchFilter
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from rest_framework.response import Response

from academic.models import Student
from administration.models import Term
from django.utils.timezone import now
from .models import (
    Receipt,
    Payment,
    ReceiptAllocation,
    PaymentAllocation,
    DebtRecord,
    PaymentRecord,
)
from .serializers import (
    ReceiptSerializer,
    PaymentSerializer,
    ReceiptAllocationSerializer,
    PaymentAllocationSerializer,
    DebtRecordSerializer,
    PaymentRecordSerializer,
    StudentDebtOverviewSerializer,
)


class PaymentAllocationListView(APIView):
    """
    Handles GET and POST requests for the list of insurances.
    """

    permission_classes = [IsAuthenticated]

    def get(self, request):
        """
        Retrieve a list of all insurances.
        """
        insurances = PaymentAllocation.objects.all()
        serializer = PaymentAllocationSerializer(insurances, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        """
        Create a new insurance record.
        """
        serializer = PaymentAllocationSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class PaymentAllocationDetailView(APIView):
    """
    Handles GET, PUT, PATCH, and DELETE requests for a single insurance.
    """

    permission_classes = [IsAuthenticated]

    def get_object(self, pk):
        return get_object_or_404(PaymentAllocation, pk=pk)

    def get(self, request, pk):
        """
        Retrieve details of a specific insurance.
        """
        print(request.data)
        insurance = self.get_object(pk)
        serializer = PaymentAllocationSerializer(insurance)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def put(self, request, pk):
        """
        Update an entire insurance record.
        """
        insurance = self.get_object(pk)
        serializer = PaymentAllocationSerializer(insurance, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def patch(self, request, pk):
        """
        Partially update an insurance record.
        """
        insurance = self.get_object(pk)
        serializer = PaymentAllocationSerializer(
            insurance, data=request.data, partial=True
        )
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk):
        """
        Delete an insurance record.
        """
        insurance = self.get_object(pk)
        insurance.delete()
        return Response(
            {"detail": "PaymentAllocation record deleted successfully"},
            status=status.HTTP_204_NO_CONTENT,
        )


class ReceiptAllocationListView(APIView):
    """
    Handles GET and POST requests for the list of insurances.
    """

    permission_classes = [IsAuthenticated]

    def get(self, request):
        """
        Retrieve a list of all insurances.
        """
        insurances = ReceiptAllocation.objects.all()
        serializer = ReceiptAllocationSerializer(insurances, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        """
        Create a new insurance record.
        """
        serializer = ReceiptAllocationSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ReceiptAllocationDetailView(APIView):
    """
    Handles GET, PUT, PATCH, and DELETE requests for a single insurance.
    """

    permission_classes = [IsAuthenticated]

    def get_object(self, pk):
        return get_object_or_404(ReceiptAllocation, pk=pk)

    def get(self, request, pk):
        """
        Retrieve details of a specific insurance.
        """
        insurance = self.get_object(pk)
        serializer = ReceiptAllocationSerializer(insurance)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def put(self, request, pk):
        """
        Update an entire insurance record.
        """
        insurance = self.get_object(pk)
        serializer = ReceiptAllocationSerializer(insurance, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def patch(self, request, pk):
        """
        Partially update an insurance record.
        """
        insurance = self.get_object(pk)
        serializer = ReceiptAllocationSerializer(
            insurance, data=request.data, partial=True
        )
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk):
        """
        Delete an insurance record.
        """
        insurance = self.get_object(pk)
        insurance.delete()
        return Response(
            {"detail": "ReceiptAllocation record deleted successfully"},
            status=status.HTTP_204_NO_CONTENT,
        )


# Receipts filter
class ReceiptFilter(FilterSet):
    from_date = DateFilter(field_name="date", lookup_expr="gte")
    to_date = DateFilter(field_name="date", lookup_expr="lte")
    class_level = CharFilter(method="filter_by_class_level")

    class Meta:
        model = Receipt
        fields = ["status", "student", "date"]

    def filter_by_class_level(self, queryset, name, value):
        return queryset.filter(student__class_level__name__iexact=value)


# Receipts List & Create View using DRF's ListCreateAPIView
class ReceiptsListView(generics.ListCreateAPIView):
    queryset = Receipt.objects.all()
    serializer_class = ReceiptSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_class = ReceiptFilter
    permission_classes = [IsAuthenticated]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        receipt = serializer.save()
        response_serializer = self.get_serializer(receipt)
        return Response(response_serializer.data, status=status.HTTP_201_CREATED)


class StudentReceiptsView(generics.ListAPIView):
    serializer_class = ReceiptSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        student_id = self.kwargs.get("student_id")
        return Receipt.objects.filter(student_id=student_id)


# Receipt Detail View (Retrieve, Update, Delete)
class ReceiptDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Receipt.objects.all()
    serializer_class = ReceiptSerializer
    permission_classes = [IsAuthenticated]

    def get_object(self):
        try:
            return super().get_object()
        except Http404:
            raise NotFound(detail="Receipt not found.", code=404)

    def update(self, request, *args, **kwargs):
        """
        Override the update method to handle custom validation or processing.
        """
        print(request.data)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)

        # Save the receipt and process student debt logic
        receipt = serializer.save()

        # Return response with detailed serialized data
        response_serializer = self.get_serializer(receipt)
        return Response(response_serializer.data, status=status.HTTP_200_OK)


# Payment List & Create View using DRF's ListCreateAPIView
class PaymentListView(generics.ListCreateAPIView):
    queryset = Payment.objects.all()
    serializer_class = PaymentSerializer
    filter_backends = [
        SearchFilter,
    ]
    filterset_fields = ["status", "date", "user"]
    permission_classes = [IsAuthenticated]

    def create(self, request, *args, **kwargs):
        """
        Override the create method to handle custom validation or processing.
        """
        print(request.data)
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        # Save the receipt and process student debt logic
        receipt = serializer.save()

        # Return response with detailed serialized data
        response_serializer = self.get_serializer(receipt)
        return Response(response_serializer.data, status=status.HTTP_201_CREATED)


# Payment Detail View (Retrieve, Update, Delete)
class PaymentDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Payment.objects.all()
    serializer_class = PaymentSerializer
    permission_classes = [IsAuthenticated]

    def get_object(self):
        try:
            return super().get_object()
        except Http404:
            raise NotFound(detail="Payment not found.", code=404)


class DebtRecordListView(generics.ListAPIView):
    queryset = DebtRecord.objects.all()
    serializer_class = DebtRecordSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ["term", "student", "is_reversed"]


class DebtRecordDetailView(generics.RetrieveAPIView):
    queryset = DebtRecord.objects.all()
    serializer_class = DebtRecordSerializer
    permission_classes = [IsAuthenticated]


class PaymentRecordListView(generics.ListAPIView):
    queryset = PaymentRecord.objects.all()
    serializer_class = PaymentRecordSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ["student", "debt_record", "processed_by"]
    search_fields = ["reference", "note"]


class PaymentRecordDetailView(generics.RetrieveAPIView):
    queryset = PaymentRecord.objects.all()
    serializer_class = PaymentRecordSerializer
    permission_classes = [IsAuthenticated]


class StudentDebtOverviewView(generics.RetrieveAPIView):
    queryset = Student.objects.all()
    serializer_class = StudentDebtOverviewSerializer
    permission_classes = [IsAuthenticated]

    def get_object(self):
        student_id = self.kwargs.get("student_id")
        try:
            return Student.objects.get(pk=student_id)
        except Student.DoesNotExist:
            raise NotFound(detail="Student not found.")


class UpdateStudentDebtView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        today = now().date()
        current_term = Term.objects.filter(
            start_date__lte=today, end_date__gte=today
        ).first()

        if not current_term:
            return Response(
                {"detail": "No active term found."}, status=status.HTTP_400_BAD_REQUEST
            )

        students = Student.objects.exclude(debt_records__term=current_term)

        for student in students:
            student.update_debt_for_term(current_term)

        return Response(
            {"detail": "Student debts updated successfully."}, status=status.HTTP_200_OK
        )


class UpdateAllUnrecordedDebtsView(APIView):
    """
    API view to update student debts for all past terms that haven't been updated.
    Ensures that no term is double-counted.
    """

    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        today = now().date()

        # Get all terms up to today (including current and past)
        terms = Term.objects.filter(end_date__lte=today).order_by("start_date")
        updated_count = 0
        skipped = 0

        for term in terms:
            students = Student.objects.exclude(debt_records__term=term)

            for student in students:
                student.update_debt_for_term(term)
                updated_count += 1

        return Response(
            {
                "detail": "Student debts updated for all missing terms.",
                "updated_count": updated_count,
            },
            status=200,
        )


class UpdatePastDebtsView(APIView):
    """
    API view to update student debts for selected past terms.
    Skips already updated records.
    """

    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        term_ids = request.data.get("term_ids", [])

        if not term_ids:
            return Response(
                {"detail": "Provide a list of term IDs to update."}, status=400
            )

        terms = Term.objects.filter(id__in=term_ids)
        if not terms.exists():
            return Response({"detail": "No valid terms found."}, status=404)

        updated = 0
        for term in terms:
            students = Student.objects.exclude(debt_records__term=term)
            for student in students:
                student.update_debt_for_term(term)
                updated += 1

        return Response(
            {
                "detail": f"Debts updated for selected terms.",
                "terms_updated": [term.name for term in terms],
                "students_updated": updated,
            },
            status=200,
        )


class ReverseStudentDebtView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        term_id = request.data.get("term_id")

        if not term_id:
            return Response(
                {"detail": "Term ID is required."}, status=status.HTTP_400_BAD_REQUEST
            )

        term = Term.objects.filter(id=term_id).first()
        if not term:
            return Response(
                {"detail": "Invalid term ID."}, status=status.HTTP_404_NOT_FOUND
            )

        students = Student.objects.filter(
            debt_records__term=term, debt_records__is_reversed=False
        )

        for student in students:
            student.reverse_debt_for_term(term)

        return Response(
            {"detail": f"Student debts for {term.name} reversed successfully."},
            status=status.HTTP_200_OK,
        )


class ReceiptBulkUploadView(APIView):
    """
    API View to handle bulk uploading of receipts from an Excel file.
    """

    def post(self, request, *args, **kwargs):
        file = request.FILES.get("file")
        if not file:
            return Response(
                {"error": "No file provided."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            columns = [
                "payer",
                "first_name",
                "middle_name",
                "last_name",
                "paid_for",
                "paid_through",
                "payment_date",
                "term",
                "amount",
            ]

            created_receipts = []
            updated_receipts = []
            skipped_receipts = []
            not_created = []

            for i, row in enumerate(
                sheet.iter_rows(min_row=2, values_only=True), start=2
            ):
                try:
                    data = dict(zip(columns, row))

                    # Normalize and extract names
                    first_name = (data.get("first_name") or "").strip().lower()
                    middle_name = (data.get("middle_name") or "").strip().lower()
                    last_name = (data.get("last_name") or "").strip().lower()
                    payer = (data.get("payer") or "").strip()
                    paid_for_name = (data.get("paid_for") or "").strip()
                    term_str = (data.get("term") or "").strip()
                    paid_through = (data.get("paid_through") or "").strip()
                    payment_date = data.get("payment_date")
                    amount = data.get("amount")

                    if not (first_name and last_name and paid_for_name and term_str):
                        raise ValueError("Missing required fields.")

                    # Lookup student
                    student = Student.objects.filter(
                        first_name__iexact=first_name,
                        middle_name__iexact=middle_name,
                        last_name__iexact=last_name,
                    ).first()
                    if not student:
                        raise ValueError("Student not found.")

                    # Determine payer
                    if not payer:
                        if student.parent_guardian:
                            payer = f"{student.parent_guardian.first_name} {student.parent_guardian.last_name}"
                        else:
                            raise ValueError(
                                "Payer not provided and student has no parent."
                            )

                    # Lookup paid_for
                    paid_for = ReceiptAllocation.objects.filter(
                        name__iexact=paid_for_name
                    ).first()
                    if not paid_for:
                        raise ValueError(
                            f"ReceiptAllocation '{paid_for_name}' not found."
                        )

                    # Lookup term by formatted string like "Term 1-2024"
                    if "-" not in term_str:
                        raise ValueError("Term must be in format 'Term 1-2024'.")
                    term_name, year = term_str.split("-")
                    term = Term.objects.filter(
                        name__iexact=term_name.strip(), academic_year__name=year.strip()
                    ).first()
                    if not term:
                        raise ValueError(f"Term '{term_str}' not found.")

                    # Check duplicate
                    existing = Receipt.objects.filter(
                        student=student,
                        paid_for=paid_for,
                        payment_date=payment_date,
                        amount=amount,
                        term=term,
                    ).first()
                    if existing:
                        skipped_receipts.append(
                            {
                                "receipt_number": existing.receipt_number,
                                "payer": payer,
                                "reason": "Duplicate receipt already exists.",
                            }
                        )
                        continue

                    # Create receipt
                    with transaction.atomic():
                        receipt = Receipt.objects.create(
                            payer=payer,
                            student=student,
                            paid_for=paid_for,
                            paid_through=paid_through,
                            payment_date=payment_date,
                            term=term,
                            amount=amount,
                            received_by=getattr(request.user, "accountant", None),
                        )
                        created_receipts.append(receipt)

                except Exception as e:
                    data["error"] = str(e)
                    not_created.append(data)

            return Response(
                {
                    "message": f"{len(created_receipts)} receipts uploaded.",
                    "updated": updated_receipts,
                    "skipped": skipped_receipts,
                    "not_created": not_created,
                },
                status=status.HTTP_201_CREATED,
            )

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


def download_receipt_template(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Receipts"

    # Headers expected by backend (now using names instead of IDs)
    headers = [
        "payer",  # Optional - fallback to student's parent
        "first_name",  # Required
        "middle_name",  # Optional
        "last_name",  # Required
        "paid_for",  # Name of ReceiptAllocation (e.g. "Tuition Fee")
        "paid_through",  # e.g. 'HATI MALIPO', 'BANK', 'M-PESA'
        "payment_date",  # Format: YYYY-MM-DD
        "term",  # Format: "Term 1-2024"
        "amount",  # Numeric
    ]

    # Example row
    example_data = [
        "Zainab Mwinami",  # payer (can be blank)
        "Zainab",  # first_name
        "Ali",  # middle_name
        "Mwinami",  # last_name
        "Tuition Fee",  # paid_for (name)
        "HATI MALIPO",  # paid_through
        "2025-07-07",  # payment_date
        "Term 1-2024",  # term
        150000,  # amount
    ]

    # Add headers with bold styling
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.font = Font(bold=True)

    # Add example data
    for col_num, value in enumerate(example_data, 1):
        ws.cell(row=2, column=col_num, value=value)

    # Return file in HTTP response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        "attachment; filename=receipt_upload_template.xlsx"
    )
    wb.save(response)
    return response
